# coding=UTF-8
import xlrd
import openpyxl
def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()
'''
    功能：
        读取指定Excel文件，返回Excel文件的数据
        
'''
def getExcelDocumentData(fileName):

    try:
        data = []
        workBook = xlrd.open_workbook(fileName)
        workSheet = workBook.sheet_by_index(0)
        if workBook.sheet_loaded(sheet_name_or_index=0):
            for row in range(workSheet.nrows):
                data.append(workSheet.row_values(rowx=row))
        return data
    except Exception as e:
        print(e)
        print("Error:打开文件失败！")












if __name__ == "__main__":
    book_name_xlsx = 'xlsx格式测试工作簿.xlsx'

    sheet_name_xlsx = 'xlsx格式测试表'

    value3 = [["姓名", "性别", "年龄", "城市", "职业"],
              ["111", "女", "66", "石家庄", "运维工程师"],
              ["222", "男", "55", "南京", "饭店老板"],
              ["333", "女", "27", "苏州", "保安"], ]

    write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, value3)
    read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)